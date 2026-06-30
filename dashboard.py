# =============================================================================
#  dashboard.py  —  your ranked daily shortlist (local web page + Excel export)
# =============================================================================
#  Opens a clean page in your browser showing leads ranked by composite score,
#  with sub-scores, suggested verdict, the 7 acquisition stages, follow-up
#  reminders, and notes. Everything you change saves straight to the database.
#  An "Export to Excel" button writes an .xlsx you can share or archive.
#
#  Built on Python's built-in web server (no heavy frameworks) + openpyxl.
#  Start it with run_dashboard.bat.
# =============================================================================

import html
import io
import json
import threading
import traceback
import webbrowser
from datetime import date, datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs, urlencode

import db
import storage

PORT = 8766
REA_URL = "https://www.realestate.com.au/rent/in-melbourne+cbd,+vic+3000/list-1"

# Background "Score now" job state (so the portal button can run scoring).
SCORE_STATE = {"running": False, "message": ""}


def _run_scoring():
    SCORE_STATE["running"] = True
    SCORE_STATE["message"] = "Scoring with location data…"
    try:
        import score_all
        score_all.main()
        SCORE_STATE["message"] = "Scoring complete."
    except Exception as e:
        SCORE_STATE["message"] = f"Scoring error: {e}"
        traceback.print_exc()
    finally:
        SCORE_STATE["running"] = False


def start_scoring():
    if SCORE_STATE["running"]:
        return False
    threading.Thread(target=_run_scoring, daemon=True).start()
    return True

VERDICT_COLOURS = {
    "Pursue & apply": "#137333",
    "Inspect first": "#b06000",
    "Pass (low priority)": "#777",
    "Pass (reject)": "#c5221f",
}
SCORE_MAX = {"score_a": 30, "score_b": 25, "score_c": 20, "score_d": 15, "score_e": 10}
SCORE_LABEL = {"score_a": "Standout", "score_b": "Location", "score_c": "Amenities",
               "score_d": "Interior", "score_e": "Pricing"}


def esc(x):
    return html.escape("" if x is None else str(x))


def effective_verdict(row):
    return row.get("verdict_override") or row.get("verdict")


def is_followup_due(row):
    fu = row.get("follow_up_date")
    if not fu:
        return False
    try:
        return datetime.fromisoformat(fu).date() <= date.today()
    except Exception:
        return False


# ---------------------------------------------------------------------------
#  HTML rendering
# ---------------------------------------------------------------------------

CSS = """
* { box-sizing: border-box; }
body { font-family: 'Segoe UI', Arial, sans-serif; margin: 0; background:#f4f5f7; color:#1a1a1a; }
header { background:#0b1f3a; color:#fff; padding:16px 24px; position:sticky; top:0; z-index:10; }
header h1 { margin:0; font-size:20px; }
header .meta { font-size:13px; opacity:.85; margin-top:4px; }
.controls { padding:12px 24px; background:#fff; border-bottom:1px solid #e0e0e0;
  display:flex; gap:10px; align-items:center; flex-wrap:wrap; }
.controls a, .controls button { text-decoration:none; padding:6px 12px; border-radius:6px;
  border:1px solid #c5ccd6; background:#fff; color:#0b1f3a; font-size:13px; cursor:pointer; }
.controls a.active { background:#0b1f3a; color:#fff; }
.wrap { padding:18px 24px; max-width:1100px; margin:0 auto; }
.card { background:#fff; border:1px solid #e2e6ea; border-radius:10px; padding:16px;
  margin-bottom:14px; box-shadow:0 1px 2px rgba(0,0,0,.04); }
.card.due { border-left:5px solid #b06000; }
.card.reject { opacity:.72; }
.top { display:flex; gap:16px; align-items:flex-start; }
.rank { font-size:13px; color:#888; min-width:30px; }
.comp { font-size:30px; font-weight:700; min-width:64px; text-align:center; }
.comp small { display:block; font-size:11px; font-weight:400; color:#888; }
.headinfo { flex:1; }
.headinfo h2 { margin:0 0 4px; font-size:16px; }
.headinfo h2 a { color:#0b1f3a; }
.sub { font-size:13px; color:#555; }
.badge { display:inline-block; padding:2px 9px; border-radius:12px; font-size:12px;
  color:#fff; font-weight:600; }
.chip { display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px;
  background:#eef1f5; color:#333; margin:2px 3px 0 0; }
.chip.flag { background:#fde7e6; color:#c5221f; font-weight:600; }
.chip.verify { background:#fff3e0; color:#b06000; }
.chip.hook { background:#e6f4ea; color:#137333; }
.bars { display:flex; gap:10px; margin:10px 0 6px; flex-wrap:wrap; }
.bar { font-size:11px; color:#555; width:120px; }
.bar .track { background:#eceff3; border-radius:4px; height:7px; margin-top:2px; }
.bar .fill { background:#3b6cc4; height:7px; border-radius:4px; }
.controls-row { display:flex; gap:14px; flex-wrap:wrap; margin-top:10px;
  padding-top:10px; border-top:1px solid #eee; align-items:flex-end; }
.field { display:flex; flex-direction:column; font-size:11px; color:#666; }
.field select, .field input, .field textarea { font-size:13px; padding:5px 7px;
  border:1px solid #c5ccd6; border-radius:6px; margin-top:3px; }
.field textarea { width:280px; height:34px; resize:vertical; }
.saved { font-size:12px; color:#137333; margin-left:8px; opacity:0; transition:opacity .2s; }
.why { font-size:12px; color:#777; margin-top:8px; }
details summary { cursor:pointer; }
.agent { font-size:12px; color:#444; margin-top:6px; }
.added { font-size:11px; color:#888; margin-top:4px; }
.toolbar { padding:10px 24px; background:#fff; border-bottom:1px solid #e0e0e0;
  display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
.toolbar .sep { width:1px; height:22px; background:#dde; margin:0 4px; }
.btn { text-decoration:none; padding:7px 13px; border-radius:6px; border:1px solid #c5ccd6;
  background:#fff; color:#0b1f3a; font-size:13px; cursor:pointer; }
.btn.primary { background:#137333; color:#fff; border-color:#137333; }
.btn.go { background:#0b1f3a; color:#fff; border-color:#0b1f3a; }
.btn.active { background:#0b1f3a; color:#fff; }
.filterbar { padding:8px 24px; background:#fbfcfe; border-bottom:1px solid #eee;
  display:flex; gap:8px; align-items:center; flex-wrap:wrap; font-size:13px; color:#555; }
.filterbar input, .filterbar select { font-size:13px; padding:4px 6px;
  border:1px solid #c5ccd6; border-radius:5px; }
.banner { margin:12px 24px 0; padding:10px 14px; border-radius:8px;
  background:#fff3e0; color:#8a5400; border:1px solid #ffe0b2; font-size:13px; }
.toast { position:fixed; bottom:18px; right:18px; background:#0b1f3a; color:#fff;
  padding:10px 16px; border-radius:8px; font-size:13px; opacity:0;
  transition:opacity .25s; pointer-events:none; }
.toast.show { opacity:1; }
"""

JS = """
function save(el){
  const body = {listing_id: el.dataset.id, field: el.dataset.field, value: el.value};
  fetch('/update', {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify(body)})
   .then(r=>r.json()).then(()=>{
      const s = document.getElementById('saved-'+el.dataset.id+'-'+el.dataset.field);
      if(s){ s.style.opacity=1; setTimeout(()=>{s.style.opacity=0;}, 1200); }
   });
}
function toast(msg){
  let t=document.getElementById('toast');
  t.textContent=msg; t.classList.add('show');
  setTimeout(()=>t.classList.remove('show'), 2600);
}
function openREA(){
  fetch('/open-rea').then(()=>toast('Opened realestate.com.au — browse to capture listings'));
}
function scoreNow(){
  fetch('/score',{method:'POST'}).then(r=>r.json()).then(j=>{
    toast(j.started ? 'Scoring started…' : 'Scoring already running…');
    pollScore();
  });
}
let _wasRunning=false;
function pollScore(){
  fetch('/score-status').then(r=>r.json()).then(j=>{
    const b=document.getElementById('scorebanner');
    if(j.running){
      _wasRunning=true;
      b.style.display='block'; b.textContent='⏳ '+(j.message||'Scoring…');
      setTimeout(pollScore, 2500);
    } else {
      if(_wasRunning){ _wasRunning=false; location.reload(); }
      else { b.style.display='none'; }
    }
  });
}
window.addEventListener('load', pollScore);
"""


def render_bar(row, key):
    val = row.get(key) or 0
    mx = SCORE_MAX[key]
    pct = int((val / mx) * 100) if mx else 0
    return (f'<div class="bar">{SCORE_LABEL[key]} {val}/{mx}'
            f'<div class="track"><div class="fill" style="width:{pct}%"></div></div></div>')


def render_select(row, field, options, current):
    opts = ""
    for o in options:
        sel = " selected" if (current or "") == o else ""
        opts += f'<option value="{esc(o)}"{sel}>{esc(o) or "(none)"}</option>'
    sid = f'{row["listing_id"]}-{field}'
    return (f'<select data-id="{esc(row["listing_id"])}" data-field="{field}" '
            f'onchange="save(this)">{opts}</select>'
            f'<span class="saved" id="saved-{sid}">saved ✓</span>')


def render_card(row, rank):
    ev = effective_verdict(row)
    colour = VERDICT_COLOURS.get(ev, "#555")
    due = is_followup_due(row)
    cls = "card"
    if due:
        cls += " due"
    if row.get("rejected"):
        cls += " reject"

    hooks = "".join(f'<span class="chip hook">{esc(h.strip())}</span>'
                    for h in (row.get("hooks") or "").split(",") if h.strip())
    flags = "".join(f'<span class="chip flag">{esc(f.strip())}</span>'
                    for f in (row.get("flags") or "").split(";") if f.strip())
    verify = "".join(f'<span class="chip verify">{esc(v.strip())}</span>'
                     for v in (row.get("verify_reasons") or "").split(";") if v.strip())

    bars = "".join(render_bar(row, k) for k in
                   ["score_a", "score_b", "score_c", "score_d", "score_e"])

    beds = row.get("bedrooms"); bath = row.get("bathrooms"); car = row.get("parking")
    rent = row.get("price_display") or (f"${row.get('price_per_week')}/wk"
                                        if row.get("price_per_week") else "?")
    sub = (f"{beds}BR &middot; {bath}bath &middot; {car} car &middot; "
           f"{esc(rent)} &middot; {esc(row.get('property_type'))} &middot; "
           f"{esc(row.get('suburb'))}")

    agent = (f"Agent: {esc(row.get('agent_name'))} &middot; "
             f"{esc(row.get('agent_phone'))} &middot; {esc(row.get('agency'))}")
    extra = []
    if row.get("available_date"):
        extra.append(esc(row["available_date"]))
    if row.get("bond_display"):
        extra.append("Bond " + esc(row["bond_display"]))
    if row.get("inspections"):
        extra.append("Inspect: " + esc(row["inspections"]))
    extra_line = " &middot; ".join(extra)

    why_bits = []
    if row.get("reject_reasons"):
        why_bits.append("Rejected: " + esc(row["reject_reasons"]))
    prov = " (provisional — run location scoring)" if row.get("composite_provisional") else ""

    verdict_opts = ["", "Pursue & apply", "Inspect first", "Pass (low priority)",
                    "Pass (reject)"]

    added = (row.get("date_scraped") or "")[:10]
    url = row.get("url") or "#"
    return f"""
    <div class="{cls}">
      <div class="top">
        <div class="rank">#{rank}</div>
        <div class="comp" style="color:{colour}">{row.get('composite',0)}{prov and '*' or ''}
          <small>/100</small></div>
        <div class="headinfo">
          <h2><a href="{esc(url)}" target="_blank">{esc(row.get('address'))}</a></h2>
          <div class="sub">{sub}</div>
          <div style="margin-top:6px">
            <span class="badge" style="background:{colour}">{esc(ev)}</span>
            {hooks}{flags}{verify}
          </div>
          <div class="bars">{bars}</div>
          <div class="agent">{agent}<br>{extra_line}</div>
          <div class="added">🗓 Added {esc(added) or '—'}</div>
          {('<div class="why">'+ ' | '.join(why_bits) +'</div>') if why_bits else ''}
        </div>
      </div>
      <div class="controls-row">
        <div class="field">Your verdict (override)
          {render_select(row, 'verdict_override', verdict_opts, row.get('verdict_override'))}</div>
        <div class="field">Stage
          {render_select(row, 'status', storage.STAGES, row.get('status'))}</div>
        <div class="field">Follow-up date
          <input type="date" data-id="{esc(row['listing_id'])}" data-field="follow_up_date"
            value="{esc(row.get('follow_up_date'))}" onchange="save(this)">
          <span class="saved" id="saved-{esc(row['listing_id'])}-follow_up_date">saved ✓</span></div>
        <div class="field">Notes
          <textarea data-id="{esc(row['listing_id'])}" data-field="notes"
            onchange="save(this)" placeholder="call notes, etc.">{esc(row.get('notes'))}</textarea>
          <span class="saved" id="saved-{esc(row['listing_id'])}-notes">saved ✓</span></div>
      </div>
    </div>
    """


def _qs(**d):
    clean = {k: v for k, v in d.items() if v}
    return ("/?" + urlencode(clean)) if clean else "/"


def render_page(view="shortlist", sort="composite", added_from="", added_to=""):
    rows = storage.all_listings(
        include_rejected=(view == "all"),
        added_from=added_from or None, added_to=added_to or None,
        order=sort)
    tally = storage.counts()
    due = [r for r in storage.all_listings() if is_followup_due(r)]

    cards = "".join(render_card(r, i) for i, r in enumerate(rows, 1)) or \
        ('<div class="card"><b>No listings match.</b><br>'
         'Click <b>Open realestate.com.au</b> above and browse some pages '
         '(they capture automatically), then click <b>Score now</b>. '
         'Or widen the date filter.</div>')

    due_html = ""
    if due:
        items = "".join(
            f'<li><a href="{esc(r.get("url") or "#")}" target="_blank">'
            f'{esc(r.get("address"))}</a> — {esc(r.get("status"))} '
            f'(due {esc(r.get("follow_up_date"))})</li>' for r in due)
        due_html = (f'<div class="card due"><b>⏰ Follow-ups due '
                    f'({len(due)})</b><ul>{items}</ul></div>')

    # tabs + export carry the current date filter
    def tab(name, label):
        active = " active" if view == name else ""
        href = _qs(view=name, sort=sort, **{"from": added_from, "to": added_to})
        return f'<a class="btn{active}" href="{href}">{label}</a>'

    exp = lambda scope: ("/export.xlsx" + "?" + urlencode(
        {k: v for k, v in {"scope": scope, "from": added_from,
                           "to": added_to}.items() if v}))

    today = date.today().isoformat()
    wk = (date.today() - timedelta(days=6)).isoformat()
    date_btn = lambda lbl, f, t: (
        f'<a class="btn{(" active" if (added_from==f and added_to==t) else "")}" '
        f'href="{_qs(view=view, sort=sort, **{"from": f, "to": t})}">{lbl}</a>')

    href_score = _qs(view=view, sort="composite",
                     **{"from": added_from, "to": added_to})
    href_date = _qs(view=view, sort="date",
                    **{"from": added_from, "to": added_to})
    sel_score = " selected" if sort != "date" else ""
    sel_date = " selected" if sort == "date" else ""
    sort_sel = (
        '<select onchange="location.href=this.value">'
        f'<option value="{href_score}"{sel_score}>Score (high&rarr;low)</option>'
        f'<option value="{href_date}"{sel_date}>Date added (newest)</option>'
        '</select>')

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>LuxeBot — Portal</title><style>{CSS}</style></head><body>
<header>
  <h1>LuxeBot — Acquisition Portal</h1>
  <div class="meta">{tally['alive']} alive &middot; {tally['total']} total &middot;
   <b>{tally['new_today']} new today</b> &middot; database: {esc(db.backend_name())}
   &middot; {esc(today)}</div>
</header>

<div class="toolbar">
  <button class="btn go" onclick="openREA()">🌐 Open realestate.com.au</button>
  <button class="btn primary" onclick="scoreNow()">📍 Score now</button>
  <span class="sep"></span>
  {tab('shortlist','Shortlist (alive)')}
  {tab('all','All (incl. rejected)')}
  <span class="sep"></span>
  <a class="btn" href="{exp('shortlist')}">⬇ Export shortlist</a>
  <a class="btn" href="{exp('all')}">⬇ Export all</a>
</div>

<div class="filterbar">
  <b>Added:</b>
  {date_btn('All','','')}
  {date_btn('Today',today,today)}
  {date_btn('Last 7 days',wk,today)}
  <span class="sep"></span>
  <form method="get" style="display:inline-flex; gap:6px; align-items:center;">
    <input type="hidden" name="view" value="{esc(view)}">
    <input type="hidden" name="sort" value="{esc(sort)}">
    from <input type="date" name="from" value="{esc(added_from)}">
    to <input type="date" name="to" value="{esc(added_to)}">
    <button class="btn" type="submit">Apply range</button>
  </form>
  <span class="sep"></span>
  Sort: {sort_sel}
</div>

<div id="scorebanner" class="banner" style="display:none;"></div>
<div class="wrap">
{due_html}
{cards}
</div>
<div id="toast" class="toast"></div>
<script>{JS}</script>
</body></html>"""


# ---------------------------------------------------------------------------
#  Excel export
# ---------------------------------------------------------------------------

# (header, row-key, width, wrap?, number-format)
EXPORT_COLS = [
    ("Composite", "composite", 11, False, "0"),
    ("Verdict", "_verdict", 18, True, None),
    ("Stage", "status", 14, False, None),
    ("Added", "date_scraped", 12, False, None),
    ("BR", "bedrooms", 5, False, None),
    ("Bath", "bathrooms", 5, False, None),
    ("Car", "parking", 5, False, None),
    ("Rent/wk", "price_per_week", 10, False, '"$"#,##0'),
    ("Type", "property_type", 12, False, None),
    ("Suburb", "suburb", 14, False, None),
    ("Address", "address", 34, True, None),
    ("Hooks", "hooks", 22, True, None),
    ("Flags", "flags", 20, True, None),
    ("Standout", "score_a", 9, False, "0"),
    ("Location", "score_b", 9, False, "0"),
    ("Amenities", "score_c", 9, False, "0"),
    ("Interior", "score_d", 9, False, "0"),
    ("Pricing", "score_e", 9, False, "0"),
    ("Agent", "agent_name", 18, True, None),
    ("Phone", "agent_phone", 14, False, None),
    ("Agency", "agency", 22, True, None),
    ("Available", "available_date", 16, False, None),
    ("Bond", "bond_display", 10, False, None),
    ("Inspections", "inspections", 26, True, None),
    ("Follow-up", "follow_up_date", 12, False, None),
    ("Notes", "notes", 32, True, None),
    ("Link", "url", 12, False, None),
]

VERDICT_FILL = {
    "Pursue & apply": ("C6EFCE", "006100"),
    "Inspect first": ("FFEB9C", "9C6500"),
    "Pass (low priority)": ("EDEDED", "555555"),
    "Pass (reject)": ("FFC7CE", "9C0006"),
}


NUMERIC_KEYS = {"composite", "bedrooms", "bathrooms", "parking",
                "price_per_week", "score_A", "score_B", "score_C",
                "score_D", "score_E"}


def _est_row_height(row, ev):
    """Estimate a row height so wrapped text isn't clipped."""
    import math
    lines = 1
    for h, key, w, wrap, fmt in EXPORT_COLS:
        if not wrap:
            continue
        if key == "_verdict":
            text = ev
        elif key == "url":
            text = "View" if row.get("url") else ""
        else:
            text = row.get(key)
        if not text:
            continue
        used = 0
        for part in str(text).split("\n"):
            used += max(1, math.ceil(len(part) / max(1, w - 1)))
        lines = max(lines, used)
    return min(150, max(18, lines * 14 + 3))


def export_xlsx(include_rejected=True, added_from=None, added_to=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    rows = storage.all_listings(include_rejected=include_rejected,
                                added_from=added_from, added_to=added_to)
    scope = "All listings" if include_rejected else "Shortlist (alive)"
    if added_from or added_to:
        scope += f" ({added_from or '…'} to {added_to or '…'})"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shortlist"
    n_cols = len(EXPORT_COLS)
    last_col = get_column_letter(n_cols)

    header_fill = PatternFill("solid", fgColor="FF0B1F3A")
    header_font = Font(bold=True, color="FFFFFFFF")
    band_fill = PatternFill("solid", fgColor="FFF3F6FB")
    title_fill = PatternFill("solid", fgColor="FFEAF0F8")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    num_align = Alignment(horizontal="center", vertical="center")
    wrap_top = Alignment(vertical="top", wrap_text=True)
    left_mid = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="FFDDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Title banner (row 1) ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(row=1, column=1,
                value=f"Liveluxe — {scope} — {date.today().isoformat()} "
                      f"— {len(rows)} listings")
    t.font = Font(bold=True, size=14, color="FF0B1F3A")
    t.fill = title_fill
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # --- Header (row 2) ---
    for ci, (h, _k, w, _wr, _f) in enumerate(EXPORT_COLS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 26

    # --- Data (row 3+) ---
    for idx, row in enumerate(rows):
        ri = idx + 3
        ev = row.get("verdict_override") or row.get("verdict")
        banded = (idx % 2 == 1)
        for ci, (h, key, w, wrap, fmt) in enumerate(EXPORT_COLS, 1):
            if key == "_verdict":
                val = ev
            elif key == "url":
                val = "View" if row.get("url") else None
            elif key == "date_scraped":
                val = (row.get("date_scraped") or "")[:10]   # date only
            else:
                val = row.get(key)
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = border
            if key in NUMERIC_KEYS:
                c.alignment = num_align
            elif wrap:
                c.alignment = wrap_top
            else:
                c.alignment = left_mid
            if fmt and isinstance(val, (int, float)):
                c.number_format = fmt
            if banded:
                c.fill = band_fill
            if key == "composite":
                c.font = Font(bold=True)
            if key == "_verdict" and ev in VERDICT_FILL:
                fg, fontc = VERDICT_FILL[ev]
                c.fill = PatternFill("solid", fgColor="FF" + fg)
                c.font = Font(color="FF" + fontc, bold=True)
            if key == "url" and row.get("url"):
                c.hyperlink = row["url"]
                c.font = Font(color="FF0563C1", underline="single")
        ws.row_dimensions[ri].height = _est_row_height(row, ev)

    # --- Freeze, filter, heat-scale ---
    ws.freeze_panes = "A3"
    if rows:
        last_row = len(rows) + 2
        ws.auto_filter.ref = f"A2:{last_col}{last_row}"
        ws.conditional_formatting.add(
            f"A3:A{last_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="FFF8696B",
                           mid_type="num", mid_value=50, mid_color="FFFFEB84",
                           end_type="num", end_value=100, end_color="FF63BE7B"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
#  HTTP server
# ---------------------------------------------------------------------------

class Handler(BaseHTTPRequestHandler):
    def _send(self, code, ctype, body, extra=None):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        for k, v in (extra or {}).items():
            self.send_header(k, v)
        self.end_headers()
        if body:
            self.wfile.write(body)

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        q = parse_qs(parsed.query)
        g = lambda name, d="": q.get(name, [d])[0]

        if path == "/open-rea":
            try:
                webbrowser.open(REA_URL)
            except Exception:
                pass
            self._send(204, "text/plain", b"")
            return

        if path == "/score-status":
            self._send(200, "application/json", json.dumps(SCORE_STATE).encode())
            return

        if path == "/export.xlsx":
            include_rejected = g("scope") == "all"
            scope = "all" if include_rejected else "shortlist"
            data = export_xlsx(include_rejected=include_rejected,
                               added_from=g("from") or None,
                               added_to=g("to") or None)
            self._send(
                200,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                data,
                {"Content-Disposition":
                 f'attachment; filename="luxebot_{scope}_{date.today().isoformat()}.xlsx"'})
            return

        view = "all" if g("view") == "all" else "shortlist"
        sort = "date" if g("sort") == "date" else "composite"
        body = render_page(view, sort, g("from"), g("to")).encode("utf-8")
        self._send(200, "text/html; charset=utf-8", body)

    def do_POST(self):
        path = urlparse(self.path).path
        if path == "/score":
            started = start_scoring()
            self._send(200, "application/json",
                       json.dumps({"started": started}).encode())
            return
        if path == "/update":
            length = int(self.headers.get("Content-Length", 0))
            try:
                data = json.loads(self.rfile.read(length))
                storage.update_user_field(data["listing_id"], data["field"],
                                          data["value"] or None)
                self._send(200, "application/json",
                           json.dumps({"ok": True}).encode())
            except Exception as e:
                self._send(500, "application/json",
                           json.dumps({"ok": False, "error": str(e)}).encode())
            return
        self._send(404, "text/plain", b"")

    def log_message(self, *a):
        pass


def main():
    storage.init_db()
    print("=" * 60)
    print("  LIVELUXE — Dashboard")
    print("=" * 60)
    print(f"  Open this in your browser:  http://localhost:{PORT}")
    print("  Leave this window open. Press Ctrl+C to stop.")
    print("=" * 60)
    ThreadingHTTPServer(("127.0.0.1", PORT), Handler).serve_forever()


if __name__ == "__main__":
    main()
